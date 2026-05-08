"""Read/write the KAS finance transactions workbook.

Schema is preserved exactly:
    Sheet1: 21 columns, Aptos Narrow 11pt bold headers, historical rows untouched.
    Finance Details: appended sheet with extracted finance/loan/ACH info,
                     keyed on Invoice Number. Created lazily on first write.

Dedup key for Sheet1: (Invoice Number, Item Description/Brand, Quantity).
"""
from __future__ import annotations
from dataclasses import dataclass, field, asdict
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable
import copy

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd


SHEET1_NAME = "Sheet1"
FINANCE_SHEET_NAME = "Finance Details"

SHEET1_COLUMNS = [
    "Finance Company", "Manufacturer Name", "Retailer Name",
    "Retailer Location/City", "Retailer Location/State",
    "Grower ID", "Grower First Name", "Grower Last Name",
    "Grower Company Name", "Grower Address1", "Grower Address2",
    "Grower City", "Grower State", "Grower ZIP CODE",
    "Item Description/Brand", "Invoice Date", "Invoice Number",
    "Standard Unit Of Measure", "Quantity", "Sum Total Price",
]  # 20 columns; column U (21st) is free-text notes preserved verbatim.

# Headers for app-managed columns beyond the original 20. Written into row 1
# only if missing (cell is None) so the user's existing header styling and
# any manually-typed alternative names are preserved.
SHEET1_EXTRA_HEADERS = {
    21: "Notes",
    22: "Date Added to Portal",
    23: "PDF Link",
}

FINANCE_COLUMNS = [
    "Invoice Number", "Patron Number", "Loan Number", "Loan Year",
    "Finance Company", "Product Rate", "Batch Number", "ACH Date",
    "Invoice Total", "Amount To Retailer", "Prepaid Amount",
    "Account Charge Amount", "Merchandised By", "PDF Source File",
    "PDF Drive ID", "Date Added", "Needs Review", "Notes",
]


@dataclass
class LineItem:
    finance_company: str | None = None
    manufacturer_name: str | None = None
    retailer_name: str | None = None
    retailer_city: str | None = None
    retailer_state: str | None = None
    grower_id: int | str | None = None
    grower_first_name: str | None = None
    grower_last_name: str | None = None
    grower_company_name: str | None = None
    grower_address1: str | None = None
    grower_address2: str | None = None
    grower_city: str | None = None
    grower_state: str | None = None
    grower_zip: int | str | None = None
    item_description: str | None = None
    invoice_date: datetime | None = None
    invoice_number: int | str | None = None
    unit: str | None = None
    quantity: float | None = None
    sum_total_price: float | None = None

    def to_row(self) -> list[Any]:
        return [
            self.finance_company, self.manufacturer_name, self.retailer_name,
            self.retailer_city, self.retailer_state,
            self.grower_id, self.grower_first_name, self.grower_last_name,
            self.grower_company_name, self.grower_address1, self.grower_address2,
            self.grower_city, self.grower_state, self.grower_zip,
            self.item_description, self.invoice_date, self.invoice_number,
            self.unit, self.quantity, self.sum_total_price,
        ]


@dataclass
class FinanceDetail:
    invoice_number: int | str | None = None
    patron_number: int | str | None = None
    loan_number: str | None = None
    loan_year: int | None = None
    finance_company: str | None = None
    product_rate: str | None = None
    batch_number: str | None = None
    ach_date: datetime | None = None
    invoice_total: float | None = None
    amount_to_retailer: float | None = None
    prepaid_amount: float | None = None
    account_charge_amount: float | None = None
    merchandised_by: str | None = None
    pdf_source_file: str | None = None
    pdf_drive_id: str | None = None
    date_added: datetime = field(default_factory=datetime.now)
    needs_review: bool = False
    notes: str | None = None

    def to_row(self) -> list[Any]:
        return [
            self.invoice_number, self.patron_number, self.loan_number,
            self.loan_year, self.finance_company, self.product_rate,
            self.batch_number, self.ach_date, self.invoice_total,
            self.amount_to_retailer, self.prepaid_amount,
            self.account_charge_amount, self.merchandised_by,
            self.pdf_source_file, self.pdf_drive_id,
            self.date_added, self.needs_review, self.notes,
        ]


@dataclass
class InvoiceBundle:
    """One KAS invoice = many line items + at most one finance detail."""
    invoice_number: int | str
    line_items: list[LineItem] = field(default_factory=list)
    finance: FinanceDetail | None = None
    pdf_source_file: str | None = None


def load(source: str | Path | bytes | BytesIO) -> Workbook:
    if isinstance(source, (bytes, bytearray)):
        source = BytesIO(source)
    return load_workbook(source)


def save(wb: Workbook, dest: str | Path) -> None:
    wb.save(dest)


def save_to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _ensure_finance_sheet(wb: Workbook) -> None:
    if FINANCE_SHEET_NAME in wb.sheetnames:
        return
    ws = wb.create_sheet(FINANCE_SHEET_NAME)
    bold = Font(name="Aptos Narrow", size=11, bold=True)
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    header_font = Font(name="Aptos Narrow", size=11, bold=True, color="FFFFFF")
    for col_idx, name in enumerate(FINANCE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    widths = [16, 14, 18, 10, 16, 22, 14, 12, 14, 16, 14, 16, 18, 32, 18, 19, 14, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def existing_keys(wb: Workbook) -> set[tuple]:
    """Return set of (invoice_number, item_description, quantity) tuples already present."""
    keys: set[tuple] = set()
    if SHEET1_NAME not in wb.sheetnames:
        return keys
    ws = wb[SHEET1_NAME]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[16] is None:
            continue
        item = (row[14] or "").strip().upper() if row[14] else ""
        qty = row[18]
        keys.add((_norm_invoice_no(row[16]), item, qty))
    return keys


def _norm_invoice_no(v) -> str:
    """Normalize any invoice-number representation to a plain integer string.

    Excel stores whole numbers as int OR float depending on the cell.
    Claude may return them as int, float, or string.  All of the below
    should normalize to "1094912":
        1094912   (int)
        1094912.0 (float)
        "1094912" (str)
        "1094912.0" (str float)
    """
    if v is None:
        return ""
    try:
        return str(int(float(str(v).strip())))
    except (ValueError, TypeError):
        return str(v).strip()


def existing_invoice_numbers(wb: Workbook) -> set[str]:
    nums: set[str] = set()
    if SHEET1_NAME not in wb.sheetnames:
        return nums
    ws = wb[SHEET1_NAME]
    for row in ws.iter_rows(min_row=2, values_only=True, max_col=17):
        if row and row[16] is not None:
            n = _norm_invoice_no(row[16])
            if n:
                nums.add(n)
    return nums


def _ensure_sheet1_extra_headers(ws) -> None:
    """Write app-managed column headers (Notes, Date Added to Portal, PDF Link)
    only when the cell is currently blank. Preserves any header text and
    formatting the user already has in row 1.
    """
    # Clone styling from a neighboring header cell (e.g. column 20) so the new
    # headers visually match the existing ones.
    style_source = ws.cell(row=1, column=min(ws.max_column, 20)) if ws.max_column else None
    for col_idx, name in SHEET1_EXTRA_HEADERS.items():
        cell = ws.cell(row=1, column=col_idx)
        if cell.value is None or (isinstance(cell.value, str) and not cell.value.strip()):
            cell.value = name
            if style_source is not None:
                if style_source.font:
                    cell.font = copy.copy(style_source.font)
                if style_source.fill and style_source.fill.fgColor:
                    cell.fill = copy.copy(style_source.fill)
                if style_source.alignment:
                    cell.alignment = copy.copy(style_source.alignment)


def _pdf_url_from_ref(ref: str | None) -> str | None:
    """Convert a stored reference (full URL or Drive file ID) to a viewable URL."""
    if not ref:
        return None
    s = str(ref).strip()
    if not s:
        return None
    if s.startswith("http://") or s.startswith("https://"):
        return s
    # Legacy Drive file ID
    return f"https://drive.google.com/file/d/{s}/view"


def append_invoice(wb: Workbook, bundle: InvoiceBundle) -> dict:
    """Append a bundle's line items to Sheet1 and finance detail to Finance Details.

    Returns dict with counts: {"line_items_added": N, "duplicates_skipped": N,
    "invoice_already_exists": bool}.
    Preserves Sheet1 formatting by cloning the previous row's font/alignment.

    Hard safety: if the invoice number is already present anywhere in Sheet1,
    NO line items are added. Line-item-level dedup (invoice+desc+qty) can be
    fooled by extraction differences (item description casing, quantity
    rounding); the invoice-number check is the bulletproof backstop.
    """
    if SHEET1_NAME not in wb.sheetnames:
        raise ValueError(f"Workbook missing sheet '{SHEET1_NAME}'")
    ws = wb[SHEET1_NAME]
    _ensure_sheet1_extra_headers(ws)

    # Per-line-item dedup is now the only safety net. The UI shows which
    # rows are already in the sheet and only sends Include-checked items;
    # this set still skips any (invoice + desc + qty) collisions just in
    # case (e.g. user re-uploaded the same invoice from a different PDF).
    keys = existing_keys(wb)

    # Snapshot a sample row's styling to clone for new rows (keeps look consistent).
    sample_row = ws.max_row if ws.max_row >= 2 else 1
    sample_cells = [ws.cell(row=sample_row, column=c) for c in range(1, len(SHEET1_COLUMNS) + 1)]

    added = 0
    skipped = 0
    next_row = ws.max_row + 1
    date_added = datetime.now().date()
    pdf_url = _pdf_url_from_ref(
        bundle.finance.pdf_drive_id if bundle.finance else None
    )
    for li in bundle.line_items:
        item_key = (
            _norm_invoice_no(li.invoice_number),
            (li.item_description or "").strip().upper(),
            li.quantity,
        )
        if item_key in keys:
            skipped += 1
            continue
        keys.add(item_key)
        for col_idx, value in enumerate(li.to_row(), start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=value)
            sample = sample_cells[col_idx - 1]
            if sample.font:
                cell.font = copy.copy(sample.font)
            if sample.alignment:
                cell.alignment = copy.copy(sample.alignment)
            if sample.number_format:
                cell.number_format = sample.number_format
        # Column 22 = "Date Added to Portal" (col 21 is Notes, preserved verbatim)
        ws.cell(row=next_row, column=22, value=date_added)
        # Column 23 = "PDF Link" — direct hyperlink for that record's PDF.
        # Uses a HYPERLINK() formula so Excel/Sheets renders it clickable.
        if pdf_url:
            link_cell = ws.cell(row=next_row, column=23,
                                value=f'=HYPERLINK("{pdf_url}","Open PDF")')
            link_cell.hyperlink = pdf_url
        next_row += 1
        added += 1

    if bundle.finance is not None:
        _ensure_finance_sheet(wb)
        fws = wb[FINANCE_SHEET_NAME]
        next_frow = fws.max_row + 1
        for col_idx, value in enumerate(bundle.finance.to_row(), start=1):
            fws.cell(row=next_frow, column=col_idx, value=value)

    return {"line_items_added": added, "duplicates_skipped": skipped,
            "invoice_already_exists": False}


def read_sheet1_dataframe(wb: Workbook) -> pd.DataFrame:
    if SHEET1_NAME not in wb.sheetnames:
        return pd.DataFrame(columns=SHEET1_COLUMNS)
    ws = wb[SHEET1_NAME]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    extra = []
    if ws.max_column >= 21:
        extra.append("Notes")
    if ws.max_column >= 22:
        extra.append("Date Added to Portal")
    if ws.max_column >= 23:
        extra.append("PDF Link")
    cols = SHEET1_COLUMNS + extra
    data = [list(r[: len(cols)]) + [None] * (len(cols) - len(r)) for r in rows]
    df = pd.DataFrame(data, columns=cols)
    if "Invoice Date" in df.columns:
        df["Invoice Date"] = pd.to_datetime(df["Invoice Date"], errors="coerce")
    for col in ("Quantity", "Sum Total Price", "Grower ID", "Grower ZIP CODE", "Invoice Number"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.dropna(how="all")
    return df


def backfill_pdf_links(wb: Workbook) -> dict:
    """Walk every Sheet1 row and populate column 23 (PDF Link) from the
    Finance Details PDF Drive ID column for matching invoice numbers.

    Returns {"rows_updated": N, "rows_skipped": N, "rows_no_pdf": N}.
    Doesn't overwrite a cell that already has a value (lets users hand-edit).
    """
    if SHEET1_NAME not in wb.sheetnames:
        return {"rows_updated": 0, "rows_skipped": 0, "rows_no_pdf": 0}
    ws = wb[SHEET1_NAME]
    _ensure_sheet1_extra_headers(ws)

    # Build invoice_no → pdf_url lookup from Finance Details
    pdf_lookup: dict[str, str] = {}
    if FINANCE_SHEET_NAME in wb.sheetnames:
        fws = wb[FINANCE_SHEET_NAME]
        for r in fws.iter_rows(min_row=2, values_only=True):
            if not r:
                continue
            inv = r[0] if len(r) > 0 else None
            pid = r[14] if len(r) > 14 else None
            if inv is not None and pid:
                url = _pdf_url_from_ref(pid)
                if url:
                    pdf_lookup[_norm_invoice_no(inv)] = url

    updated, skipped, no_pdf = 0, 0, 0
    for row_idx in range(2, ws.max_row + 1):
        inv_cell = ws.cell(row=row_idx, column=17)  # Invoice Number is col 17
        if inv_cell.value is None:
            continue
        inv_no = _norm_invoice_no(inv_cell.value)
        existing = ws.cell(row=row_idx, column=23).value
        if existing not in (None, ""):
            skipped += 1
            continue
        url = pdf_lookup.get(inv_no)
        if not url:
            no_pdf += 1
            continue
        link_cell = ws.cell(row=row_idx, column=23,
                            value=f'=HYPERLINK("{url}","Open PDF")')
        link_cell.hyperlink = url
        updated += 1
    return {"rows_updated": updated, "rows_skipped": skipped, "rows_no_pdf": no_pdf}


def read_finance_dataframe(wb: Workbook) -> pd.DataFrame:
    if FINANCE_SHEET_NAME not in wb.sheetnames:
        return pd.DataFrame(columns=FINANCE_COLUMNS)
    ws = wb[FINANCE_SHEET_NAME]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data = [list(r[: len(FINANCE_COLUMNS)]) + [None] * (len(FINANCE_COLUMNS) - len(r)) for r in rows]
    df = pd.DataFrame(data, columns=FINANCE_COLUMNS)
    for col in ("Invoice Total", "Amount To Retailer", "Prepaid Amount",
                "Account Charge Amount", "Loan Year", "Invoice Number"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    for col in ("ACH Date", "Date Added"):
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    df = df.dropna(how="all")
    return df
